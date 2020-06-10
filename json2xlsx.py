import configparser
import os
import json
import cv2
import re
import requests
import pytesseract
import io
import shutil
import hashlib
import logging
import numpy as np

from openpyxl import Workbook
from PIL import Image
from urllib.request import urlopen
from urllib import parse

# Config Parser 초기화
config = configparser.ConfigParser()
# Config File 읽기
config.read(os.path.dirname(os.path.realpath(__file__)) + os.sep + 'envs' + os.sep + 'property.ini')

logging.basicConfig(filename='json2xlsx.log', level=logging.INFO)

def getFiles(path: str, files: list):
    listdirs = os.listdir(path)

    for listdir in listdirs:
        subpath = path + '/' + listdir
        # rint(subpath)
        if os.path.isdir(subpath):
            getFiles(subpath, files)
        else:
            files.append(subpath)

def md5Generator(text: str):
    enc = hashlib.md5()
    enc.update(text.encode('utf-8'))
    encText = enc.hexdigest()
    return encText

def url_to_image(url, readFlog=cv2.IMREAD_COLOR):
    p = re.compile('[가-힣]+')
    hangeuls = p.findall(url)

    for hangeul in hangeuls:
        url = url.replace(hangeul, parse.quote(hangeul))

    resp = urlopen(url)
    image = np.asarray(bytearray(resp.read()), dtype=np.uint8)
    image = cv2.imdecode(image, readFlog)

    return image

# 이미지 -> 문자열 추출
def ocrToStr(img, lang='kor+eng'):
    # 추출(이미지파일, 추출언어, 옵션)
    config_custom = '--psm 1 -c preserve_interword_spaces=1'
    ocrText = pytesseract.image_to_string(img, lang=lang, config=config_custom)

    outText = ocrText.replace('\t', '').replace('\n', ' ').replace('  ', '')
    return outText

# 문자열 -> 텍스트파일 개별 저장
def strToTxt(txtName, outText):
    with open(txtName + '.txt', 'at', encoding='utf-8') as f:
        f.write(outText)
        f.write(',')

def main():
    # json 파일 경로
    jsonPath = os.path.dirname(os.path.realpath(__file__)) + config['Path']['JsonPath']

    # json to ocr 완료된 파일 경로
    jsonOkPath = os.path.dirname(os.path.realpath(__file__)) + config['Path']['JsonOkPath']

    # xlsx 파일 경로
    xlsxPath = os.path.dirname(os.path.realpath(__file__)) + config['Path']['XlsxPath']

    file_seq = 1
    xlsx_seq = 1

    logging.info("Step #1. json 파일 전체를 배열에 담기 ")
    logging.info(" ### %s " % jsonPath)
    jsonFiles = []
    getFiles(jsonPath, jsonFiles)
    logging.info(" ### 총 건수 %d " % len(jsonFiles))

    logging.info("Step #2. 엑셀 파일 만들기 %d" % xlsx_seq)
    row = 1
    wb = Workbook()
    ws = wb.active
    ws.cell(row, column=1).value = 'doc_id'
    ws.cell(row, column=2).value = 'doc_title'
    ws.cell(row, column=3).value = 'doc_datetime'
    ws.cell(row, column=4).value = 'doc_url'
    ws.cell(row, column=5).value = 'img_url'
    ws.cell(row, column=6).value = 'ocr_result'
    ws.cell(row, column=7).value = 'ocr_yn'

    for jsonFile in jsonFiles:
        with open(jsonFile) as jsonF:
            logging.info("Step #3. json 파일 읽기 %d. %s" % (file_seq, jsonFile))
            jsonData = json.load(jsonF)

            for imgUrl in jsonData['img_url']:
                logging.info("Step #4. 이미지 URL -> OCR %d" % row)
                logging.info("Step #4-1. %s " % imgUrl)

                try:
                    # response = requests.get(imgUrl)
                    # img = Image.open(io.BytesIO(response.content))

                    img = url_to_image(imgUrl)
                    ocrResult = ocrToStr(img, 'kor+eng')
                    ocr_yn = 'N'

                    logging.info("Step #4-2. %s " % ocrResult.strip())

                    if len(ocrResult.strip()) > 5:
                        ocr_yn = 'Y'

                    logging.info("Step #4-3. %s " % ocr_yn)

                    row += 1
                    doc_id = md5Generator(jsonData['doc_url'])

                    ws.cell(row, column=1).value = doc_id
                    ws.cell(row, column=2).value = jsonData['doc_title']
                    ws.cell(row, column=3).value = jsonData['doc_datetime']
                    ws.cell(row, column=4).value = jsonData['doc_url']
                    ws.cell(row, column=5).value = imgUrl
                    ws.cell(row, column=6).value = ocrResult.strip()
                    ws.cell(row, column=7).value = ocr_yn

                except:
                    pass


        logging.info("Step #5. json 파일 이동 %d. %s" % (file_seq, jsonOkPath))
        shutil.move(jsonFile, jsonOkPath)

        if (file_seq > 1000):
            filename = 'result_' + str(xlsx_seq)
            wb.save(xlsxPath + '/' + filename + '.xlsx')
            logging.info("Step #6. 엑셀 파일 저장: %s " % (xlsxPath + '/' + filename + '.xlsx'))

            file_seq = 1
            xlsx_seq += 1
            logging.info("Step #2. 엑셀 파일 만들기 %d" % xlsx_seq)
            row = 1
            wb = Workbook()
            ws = wb.active
            ws.cell(row, column=1).value = 'doc_id'
            ws.cell(row, column=2).value = 'doc_title'
            ws.cell(row, column=3).value = 'doc_datetime'
            ws.cell(row, column=4).value = 'doc_url'
            ws.cell(row, column=5).value = 'img_url'
            ws.cell(row, column=6).value = 'ocr_result'
            ws.cell(row, column=7).value = 'ocr_yn'
            break
        else:
            file_seq += 1

    filename = 'result_' + str(xlsx_seq)
    wb.save(xlsxPath + '/' + filename)
    logging.info("Step #6. 엑셀 파일 저장: %s " % (xlsxPath + '/' + filename ))
    logging.info("Step #7. 끗")

if __name__ == "__main__":
    main()